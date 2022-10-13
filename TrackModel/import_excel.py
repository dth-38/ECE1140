import pandas as pd
from openpyxl import load_workbook

def get_sheets(filepath):
    ## Get all sheet names from the workbook
    wb = load_workbook(filepath, read_only = True, keep_links = False)
    sheets = wb.sheetnames
    
    ## Get sheets with individual line information
    line_sheets = []
    for i in range(len(sheets)):
        s = sheets[i]
        if s.find('Line') != -1:
            line_sheets.append(sheets[i])

    ## Return the names of sheets with line info
    return line_sheets

sht = get_sheets('track_layout.xlsx')
print(sht)