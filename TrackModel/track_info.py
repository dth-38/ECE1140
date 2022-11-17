import sys
import pandas as pd
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem
from openpyxl import load_workbook

class TrackInfo:
    def __init__(self, *, fp=''):
        self.filepath = fp
        self.sheets = self.propogate_sheets()
        self.red_table = QTableWidget()
        self.green_table = QTableWidget()
        self.blue_table = QTableWidget()
        self.current_table = QTableWidget()
        self.red_length = 0
        self.green_length = 0

    def set_filepath(self, fp):
        self.filepath = fp
        self.sheets = self.propogate_sheets()

    def get_filepath(self):
        return self.filepath

    def get_sheet(self, num):
        return self.sheets[num]

    def get_length(self, line):
        if (line.lower() == "red"):
            return self.red_length
        elif (line.lower() == "green"):
            return self.green_length

    def get_cell_text(self, table, row, col):
        ## Return value of table cell
        cell = self.get_cell(table, row, col)
        return cell.text()

    def get_cell(self, table, row, col):
        cell = table.item(row, col)
        return cell

    def set_dimensions(self):
        ## Iterate through red and green table, add to length variable
        #  for each if in the correct table
        r = 0
        if (self.red_table.rowCount() >= self.green_table.rowCount()):
            r = self.red_table.rowCount()
        else:
            r = self.green_table.rowCount()

        for x in range(r - 1):
            x = x + 1
            red_cell = self.get_cell(self.red_table, x, 0)
            green_cell = self.get_cell(self.green_table, x, 0)
            
            ## If cell is blank (NoneType), set text to ""
            if red_cell is None:
                self.red_table.setRowCount(x + 1)
                red_cell = QTableWidgetItem()
                red_cell.setText("")
                self.red_table.setItem(x, 0, red_cell)
            if green_cell is None:
                self.green_table.setRowCount(x + 1)
                green_cell = QTableWidgetItem()
                green_cell.setText("")
                self.green_table.setItem(x, 0, green_cell)
                
            ## Add to table length if cell is filled
            if (red_cell.text().lower() == "red"):
                self.red_length += 1
            if (green_cell.text().lower() == "green"):
                self.green_length += 1

        ## Set new row counts
        self.red_table.setRowCount(self.red_length)
        self.green_table.setRowCount(self.green_length)

    def propogate_sheets(self):
        ## Get all sheet names from the workbook
        wb = load_workbook(self.filepath, read_only = True, keep_links = False)
        sh = wb.sheetnames
    
        ## Get sheets with individual line information
        line_sheets = []
        for i in range(len(sh)):
            s = sh[i]
            if s.find('Line') != -1:
                line_sheets.append(sh[i])

        ## Fill sheets[] with the names of sheets with line info
        return line_sheets

    ## Place excel data in PyQt5 table
    def load_excel_data(self, worksheet_name, table):
        ## Fill out red_table and green_table when load function is called
        #  saves values from table in local variables
        if (worksheet_name.lower() == "red line"):
            self.red_table = table
        elif (worksheet_name.lower() == "green line"):
            self.green_table = table
        elif (worksheet_name.lower() == "blue line"):
            self.blue_table = table
        
        df = pd.read_excel(self.filepath, worksheet_name)
        if df.size == 0:
            return

        df.fillna('', inplace=True)
        table.setRowCount(df.shape[0])
        table.setColumnCount(df.shape[1])
        table.setHorizontalHeaderLabels(df.columns)

        # returns pandas array object
        for row in df.iterrows():
            values = row[1]
            for col_index, value in enumerate(values):
                if isinstance(value, (float, int)):
                    value = '{0:0,.0f}'.format(value)
                tableItem = QTableWidgetItem(str(value))
                table.setItem(row[0], col_index, tableItem)

        table.setColumnWidth(2, 300)
        return table
    